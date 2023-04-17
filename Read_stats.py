import pandas as pd
import re
import openpyxl
import sys
import os
#setting path
path=input("Give the path of directory having the raw and HQ files=")
os.chdir(path)
#opening and reading file and converting from txt to xlsx

#raw file and selecting the required columns
raw_file= input("Enter raw file name to proceed=")
File_raw = pd.read_table(raw_file, delimiter = "\t" )
File_raw.to_excel('File_raw.xlsx', index=False, header=True)
raw= pd.read_excel('File_raw.xlsx', header=0, usecols="A,D,E", names=['Sample_Name', 'num_seqs_raw', 'sum_len_raw'])

#Hq file and selecting the required columns
HQ_file= input("Enter HQ file name to proceed=")
File_HQ = pd.read_table(HQ_file, delimiter = "\t" )
File_HQ.to_excel('File_HQ.xlsx', index=False, header=True)
HQ= pd.read_excel('File_HQ.xlsx', header=0, usecols="A,D,E", names=['Sample_Name_HQ', 'num_seqs_HQ', 'sum_len_HQ'])

#concatenating Raw and HQ file 
File_1=pd.concat([raw, HQ], axis=1)
File_1.to_excel('File_1.xlsx', index=False, header=True)

#opening the newly created excel file with raw and HQ stats
wb_temp = openpyxl.load_workbook('File_1.xlsx') 
sheet_temp = wb_temp.active

#removing the extensions from the sample name 
for i in range(2, sheet_temp.max_row+1): 
    a=sheet_temp.cell(row=i, column=1).value            
    b=sheet_temp.cell(row=i, column=4).value            
    val_1 = re.sub(r"_001.fastq.gz$","", a)
    val_1_2= re.sub(r"_S[0-9]*","",val_1)
    val_1_3= re.sub(r"_L[0-9]*","",val_1_2)          #change according to the sample received 
    val_4= re.sub(r".fq$","", b)
    val_4_2= re.sub(r"_S[0-9]*","",val_4)
    sheet_temp.cell(row=i, column=1).value = val_1_3
    sheet_temp.cell(row=i, column=4).value = val_4_2
wb_temp.save("File_1.xlsx")

#Splitting the concatenated file and sorting the raw data and Hq data
File_2=pd.read_excel('File_1.xlsx')
raw_sorted=File_2.loc[:,['Sample_Name', 'num_seqs_raw','sum_len_raw']]
raw_sorted.sort_values(by=['Sample_Name'],ascending=True, inplace=True,ignore_index=True )
HQ_sorted= File_2.loc[:,['Sample_Name_HQ', 'num_seqs_HQ','sum_len_HQ']]
HQ_sorted.sort_values(by=['Sample_Name_HQ'],ascending=True,inplace=True,ignore_index=True ) #inplace=True makes sure the changes are made to existing dataframe

File=pd.concat([raw_sorted, HQ_sorted], axis=1)
File.to_excel('File.xlsx', index=False, header=True)
#opening the file for further process
wb= openpyxl.load_workbook('File.xlsx') 
sheet = wb.active

#checking if sample names of both the files are matching 
check_point1=0
for i in range(2, sheet.max_row+1):   
     if sheet.cell(row=i, column=1).value==sheet.cell(row=i, column=4).value:
         print("Sample name matching")
     else:
         print(f"Sample name not matching at row no:{str(i)}")
         check_point1+=1
#check point
sys.exit("Error in data. Program end.") if check_point1>=1 else print("Ready to proceed")

#deleting HQ name column
sheet.delete_cols(4)
wb.save("File.xlsx")

#calculate total bases
def total_bases(R1,R2):
    global result
    result=[]
    tup= tuple(zip(R1, R2))
    for i in tup:
        s=sum(i)
        result.append(s)  
    return result

#calculate total bases in GB
def bases_in_GB(tb):
    global gb
    gb=[]
    for i in tb:
        tb_gb=round(i/1000000000, 2)
        gb.append(tb_gb)  
    return gb

#read count check of R1 and R2
def read_count_check(a,b):
    tup=tuple(zip(a,b))
    global reads
    reads=[]
    for i, j in tup:
        if i==j:
            reads.append(i)
        else:
            print("Not matching reads") 
    return reads


Sample_name= File['Sample_Name'].values.tolist()
R1_list=[]
R2_list=[]
ind1=[]
ind2=[]
Read_count_raw=File['num_seqs_raw'].values.tolist()
Bases_count_raw=File['sum_len_raw'].values.tolist()
Read_count_HQ=File['num_seqs_HQ'].values.tolist()
Bases_count_HQ=File['sum_len_HQ'].values.tolist()

for i in Sample_name:
    if i.endswith("R1"):
        ind1.append(Sample_name.index(i))
        R1_list.append(i)
    elif i.endswith("R2"):
        ind2.append(Sample_name.index(i))
        R2_list.append(i)

#creating lists to store each column elements required to make the final sheet
R1_reads_raw=[]
R1_bases_raw=[]
R1_reads_HQ=[]
R1_bases_HQ=[]
R2_reads_raw=[]
R2_bases_raw=[]
R2_reads_HQ=[]
R2_bases_HQ=[]
for a in ind1:
    R1_reads_raw.append(Read_count_raw[a])
    R1_bases_raw.append(Bases_count_raw[a])
    R1_reads_HQ.append(Read_count_HQ[a])
    R1_bases_HQ.append(Bases_count_HQ[a])
for b in ind2:
    R2_reads_raw.append(Read_count_raw[b])
    R2_bases_raw.append(Bases_count_raw[b])
    R2_reads_HQ.append(Read_count_HQ[b])
    R2_bases_HQ.append(Bases_count_HQ[b])

totalbases_raw= total_bases(R1_bases_raw,R2_bases_raw)
basesGB_raw= bases_in_GB(totalbases_raw)
totalbases_HQ= total_bases(R1_bases_HQ,R2_bases_HQ)
basesGB_HQ= bases_in_GB(totalbases_HQ)

#getting the name of the samples and removing "_R1" and "_R2" from it
names=[]     
for i in ind1:
    temp= Sample_name[i]
    names.append(re.sub(r"_R1","", temp))

#creating dictionary for dataframe
mydataset = {
  'Sample Name': names,
  'R1_reads_raw':read_count_check(R1_reads_raw, R2_reads_raw),
  'R1_bases_raw': R1_bases_raw,
  'R2_bases_raw': R2_bases_raw,
  'Total_bases_raw ': totalbases_raw,
  'Raw_data_in_GB':basesGB_raw,
  'R1_reads_HQ':read_count_check(R1_reads_HQ, R2_reads_HQ),
  'R1_bases_HQ':R1_bases_HQ,
  'R2_bases_HQ':R2_bases_HQ,
  'Total_bases_HQ':totalbases_HQ,
  'HQ_data_in_GB':basesGB_HQ
}
#final read_stats sheet
Final_result = pd.DataFrame(mydataset)
#sorting sheet in ascending order
Final_result.sort_values(by=["Sample Name"], inplace=True ) 

#dataframe to excel sheet 
Final_result.to_excel('Read_stats.xlsx', sheet_name='Sheet1', columns=None, index=False)
#loading read stats sheet
wb_reads = openpyxl.load_workbook('Read_stats.xlsx') 
sheet_read= wb_reads.active

                                              #####comparison with sample sheet#####

SS= input("Enter Sample sheet name to proceed=")
Sample_sheet= pd.read_excel(SS)
Sample_sheet.sort_values(by=['Sample Name In Sample Sheet'], inplace=True ) #sorting sheet in ascending order
Sample_sheet.to_excel('Sample_sheet_sorted.xlsx', index=False, header=True)

#loading sample sheet
wb_sample = openpyxl.load_workbook("Sample_sheet_sorted.xlsx") 
sheet_sample = wb_sample.active

#checking names
check_point2=0
for i in range(2, sheet_sample.max_row+1): 
    a=sheet_read.cell(row=i, column=1).value #read stats
    b=sheet_sample.cell(row=i, column=5).value #Sample sheet
    if a==b:
        print(f"Names are matching in sample sheet and read stats sheet for sample={str(a)}") 
        
    else:
        print(f"Names are not matching for sample={str(a)}")
        check_point2+=1
#check point
sys.exit("Error in data. Program end.") if check_point2>=1 else print("Ready to proceed")

#adding columns from the sample sheet
Read_stats= pd.read_excel('Read_stats.xlsx', header=0)
ProjectID_clientname=pd.read_excel('Sample_sheet_sorted.xlsx', header=0, usecols="B,C")
temp1=pd.concat([ProjectID_clientname,Read_stats], axis=1)
data_comm= pd.read_excel('Sample_sheet_sorted.xlsx', header=0, usecols="K", names=['Data committed in GB'])
temp2=pd.concat([temp1, data_comm], axis=1)
temp2.to_excel('Read_stats_temp.xlsx',index=True, header=True)

#open the temp file and add the extra new columns
wb_RST= openpyxl.load_workbook("Read_stats_temp.xlsx") 
sheet_RST = wb_RST.active
sheet_RST['P1']="Data difference"
sheet_RST['Q1']="Pass/Fail"
sheet_RST['A1']="Sl.no."

#Calculating if there is enough data or not
diff=0.0
for i in range(2, sheet_RST.max_row+1):
    collect=sheet_RST.cell(row=i, column=14).value 
    commit=sheet_RST.cell(row=i, column=15).value 
    #print(collect,commit)
    diff= float(collect-commit)
    sheet_RST.cell(row=i, column=16).value=diff
    if diff>=0:
        sheet_RST.cell(row=i, column=17).value= "Pass"
    else:
        sheet_RST.cell(row=i, column=17).value= "Fail"

#Correcting the index
inc=1
for i in range(2, sheet_RST.max_row+1):
    sheet_RST.cell(row=i, column=1).value=inc
    inc+=1
wb_RST.save("Read_stats_temp.xlsx")

#Adding the final columns to the sheet
Final_cols= pd.read_excel('Sample_sheet_sorted.xlsx', header=0, usecols="L,M,N")
last=pd.read_excel('Read_stats_temp.xlsx', header=0)
temp3=pd.concat([last, Final_cols], axis=1)
temp3.to_excel('FINAL_read_stats.xlsx',index=False, header=True)

wb_FRS= openpyxl.load_workbook("FINAL_read_stats.xlsx") 
sheet_FRS = wb_FRS.active

#adding colours to the read_stats sheet columns
from openpyxl.styles import PatternFill
filler_1=PatternFill(patternType='solid', fgColor='fa4c47') #hex code for colours
filler_2=PatternFill(patternType='solid', fgColor='504cd1')    
cell_ids_raw = ['E1', 'F1', 'G1', 'H1', 'I1']
cell_ids_HQ = ['J1', 'K1', 'L1', 'M1', 'N1']
for i in range(5):
    sheet_FRS[cell_ids_raw[i]].fill = filler_1
    sheet_FRS[cell_ids_HQ[i]].fill = filler_2
wb_FRS.save("FINAL_read_stats.xlsx")

os.remove('File.xlsx')
os.remove('Read_stats_temp.xlsx')
os.remove('File_raw.xlsx')
os.remove('File_HQ.xlsx')